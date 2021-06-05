from openpyxl import load_workbook
from datetime import datetime

NOT_PLANNED = 0
DONE = 1
NOT_DONE = 2


class ExcelMaker:
    @staticmethod
    def convert_in_excel(tasks_list, filename):
        wb = load_workbook(filename='template.xlsx')
        weekly_data = ExcelMaker.get_weekly_data(tasks_list)
        ExcelMaker.write_weekly_data(wb, weekly_data)
        ExcelMaker.write_last_weekly_data(wb, tasks_list)

        wb.save(filename)

    @staticmethod
    def write_last_weekly_data(wb, tasks_list):
        last_week_data = sorted(tasks_list, key=lambda x: datetime.strptime(x['start_date'], '%d.%m.%Y'), reverse=True)[0]
        sheet = ExcelMaker.get_sheet_by_name(wb, 'trello_last_week_data')

        row_index = 2
        actions = last_week_data['actions']
        for activity_name in actions:
            sheet.cell(row=row_index, column=1).value = activity_name
            column_index = 2
            action = actions[activity_name]
            for day_of_week in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']:
                is_Planned = day_of_week in action['plan']
                is_Done = day_of_week in action['fact']

                done_index = -1
                if is_Done:
                    done_index = DONE
                elif is_Planned:
                    done_index = NOT_DONE
                else:
                    done_index = NOT_PLANNED

                sheet.cell(row=row_index, column=column_index).value = done_index

                column_index += 1

            row_index += 1

    @staticmethod
    def write_weekly_data(wb, weekly_data):
        weekly_worksheet = ExcelMaker.get_sheet_by_name(wb, 'trello_weekly_data')

        row_index = 2
        date_row_index = 1
        date_written = False
        for activity_name in weekly_data:
            activity = weekly_data[activity_name]
            column_index = 1
            weekly_worksheet.cell(row=row_index, column=column_index).value = activity_name

            for start_date in sorted([key for key in activity.keys()], key=lambda x: datetime.strptime(x, '%d.%m.%Y')):
                weekly_activity_data = activity[start_date]
                weekly_worksheet.cell(row=row_index, column=column_index + 1).value = weekly_activity_data['fact']
                weekly_worksheet.cell(row=row_index, column=column_index + 2).value = weekly_activity_data['plan']

                if not date_written:
                    weekly_worksheet.cell(row=date_row_index, column=column_index + 1).value = start_date
                    weekly_worksheet.cell(row=date_row_index, column=column_index + 2).value = weekly_activity_data['end_date']

                column_index += 2

            date_written = True
            row_index += 1

    @staticmethod
    def get_weekly_data(tasks_list):
        activities = set([actions for week in tasks_list for actions in week['actions']])
        weekly_data = {}
        for activity in activities:
            weekly_data[activity] = {}
        for week in tasks_list:
            start_date = week['start_date']
            for activity in weekly_data:
                weekly_data[activity][start_date] = {
                    "end_date": week['end_date'],
                    "plan": 0,
                    "fact": 0
                }

            actions = week['actions']

            for activity in actions:
                weekly_data[activity][start_date]['plan'] = len(actions[activity]['plan'])
                weekly_data[activity][start_date]['fact'] = len(actions[activity]['fact'])

        return weekly_data

    @staticmethod
    def get_sheet_by_name(wb, name):
        return [sheet for sheet in wb.worksheets if sheet._WorkbookChild__title == name][0]